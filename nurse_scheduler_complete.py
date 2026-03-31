import sys
import traceback
import re

try:
    import os
    import math
    import calendar
    from collections import defaultdict
    from datetime import datetime, date

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment
    from ortools.sat.python import cp_model

    SHIFT_CODES = ["D", "E", "N", "M", "O"]
    WORK_CODES = ["D", "E", "N", "M"]
    PREF_MARKERS = {"D", "E", "N", "M"}

    GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")
    YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
    GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
    BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")
    PINK_FILL = PatternFill("solid", fgColor="FCE4D6")
    RED_FILL = PatternFill("solid", fgColor="F4CCCC")

    def get_days_in_month(year, month):
        return calendar.monthrange(year, month)[1]

    def parse_date_cell(val):
        if val is None or val == "":
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        raise ValueError(f"Unsupported date value: {val!r}")

    def normalize_text(v):
        if v is None:
            return ""
        return str(v).strip().upper()

    def read_setup(ws):
        cfg = {
            "year": int(ws["B2"].value),
            "month": int(ws["B3"].value),
            "base_offs": int(ws["B4"].value) if ws["B4"].value is not None else 11,
            
            "weekday_req": {
                "D": int(ws["B6"].value),
                "E": int(ws["B7"].value),
                "N": int(ws["B8"].value),
                "M": int(ws["B9"].value) if ws["B9"].value is not None else 0,
            },
            
            "weekend_req": {
                "D": int(ws["B11"].value),
                "E": int(ws["B12"].value),
                "N": int(ws["B13"].value),
                "M": int(ws["B14"].value) if ws["B14"].value is not None else 0,
            },
            
            "max_work_run": int(ws["B16"].value) if ws["B16"].value is not None else 5,
            "max_consecutive_d": int(ws["B17"].value) if ws["B17"].value is not None else 4,
            "max_consecutive_e": int(ws["B18"].value) if ws["B18"].value is not None else 4,
            "max_night_run": int(ws["B19"].value) if ws["B19"].value is not None else 3,
            "min_night_run": int(ws["B20"].value) if ws["B20"].value is not None else 1,
            "post_night_off": int(ws["B21"].value) if ws["B21"].value is not None else 2,
            
            "time_limit": int(ws["B23"].value) if ws["B23"].value is not None else 300,
            "pref_weight": int(ws["B24"].value) if ws["B24"].value is not None else 100,
            "fair_total_weight": int(ws["B25"].value) if ws["B25"].value is not None else 10,
            "fair_night_weight": int(ws["B26"].value) if ws["B26"].value is not None else 10,
            "fair_day_weight": int(ws["B27"].value) if ws["B27"].value is not None else 4,
            "fair_evening_weight": int(ws["B28"].value) if ws["B28"].value is not None else 10,
            "fair_weekend_weight": int(ws["B29"].value) if ws["B29"].value is not None else 2,
        }
        return cfg

    def load_inputs(path):
        wb_data = load_workbook(path, data_only=True) 
        wb = load_workbook(path) 
        
        cfg = read_setup(wb_data["Setup"])
        days = get_days_in_month(cfg["year"], cfg["month"])

        nurses = []
        is_senior = []
        is_nk = []
        max_offs = []
        raw_preceptors = []
        is_ghost = []  
        is_junior = []  # 🌟 막내 여부를 담을 리스트 추가
        nurse_index = {}
        ws_nurses = wb_data["Nurses"]
        r = 2
        
        while True:
            name = ws_nurses.cell(r, 1).value
            if name is None or str(name).strip() == "":
                break
            name = str(name).strip()

            senior_val = str(ws_nurses.cell(r, 2).value).strip().upper() if ws_nurses.cell(r, 2).value else ""
            if senior_val == "Y":
                sen_shifts = set(WORK_CODES)
            elif "차지" in senior_val or "나이트" in senior_val or "전담" in senior_val:
                sen_shifts = set()
                if "D" in senior_val or "데이" in senior_val: sen_shifts.add("D")
                if "E" in senior_val or "이브" in senior_val: sen_shifts.add("E")
                if "N" in senior_val or "나이트" in senior_val: sen_shifts.add("N")
                if "M" in senior_val or "미들" in senior_val: sen_shifts.add("M")
                
                if not sen_shifts:
                    sen_shifts.add("N")
            else:
                sen_shifts = set()

            mo_val = ws_nurses.cell(r, 3).value
            try:
                mo = int(mo_val)
            except (ValueError, TypeError):
                mo = days
                
            nk_val = ws_nurses.cell(r, 4).value
            nk = True if str(nk_val).strip().upper() == "Y" else False
            
            prec_val = ws_nurses.cell(r, 5).value
            if prec_val is None:
                prec_val = ""
            prec_str = str(prec_val).strip()

            ghost_flag = False
            if "NODUTY" in prec_str.upper().replace(" ", ""):
                ghost_flag = True
                prec_str = re.sub(r'(?i)\(?\s*no\s*duty\s*\)?', '', prec_str).strip()

            # 🌟 엑셀 6번째 열(프리셉터 옆칸)에서 막내(Y/N) 값 읽어오기
            jun_val = ws_nurses.cell(r, 6).value
            jun_str = str(jun_val).strip().upper() if jun_val else "N"
            is_junior.append(jun_str)

            raw_preceptors.append(prec_str)
            is_ghost.append(ghost_flag)

            nurse_index[name] = len(nurses)
            nurses.append(name)
            is_senior.append(sen_shifts)
            is_nk.append(nk)
            max_offs.append(mo)
            r += 1

        preceptors = []
        for rp in raw_preceptors:
            if rp and rp.lower() not in ["none", ""]:
                names = rp.replace(",", " ").replace("/", " ").split()
                valid_preceptors = [nurse_index[nm] for nm in names if nm in nurse_index]
                preceptors.append(valid_preceptors)
            else:
                preceptors.append([])

        holidays = set()
        extra_reqs = defaultdict(lambda: defaultdict(int))
        
        ws_h = wb_data["Holidays"]
        for r in range(2, ws_h.max_row + 1):
            raw = ws_h.cell(r, 1).value
            if raw in (None, ""):
                continue
            dt = parse_date_cell(raw)
            if dt.year != cfg["year"] or dt.month != cfg["month"]:
                raise ValueError(f"Holidays sheet contains {dt.isoformat()}, outside selected month.")

            note = str(ws_h.cell(r, 2).value).strip().upper() if ws_h.cell(r, 2).value else ""
            shift_patterns = re.findall(r"\b([DENM])(\d*)\b", note)
            
            if shift_patterns:
                for shift, cnt_str in shift_patterns:
                    cnt = int(cnt_str) if cnt_str else 1
                    extra_reqs[dt.day - 1][shift] += cnt
                pure_shifts = re.sub(r"\b[DENM]\d*\b", "", note).replace(",", "").replace("/", "").replace(" ", "")
                if pure_shifts != "":
                    holidays.add(dt.day - 1)
            else:
                holidays.add(dt.day - 1)

        cfg["extra_reqs"] = dict(extra_reqs)
        cfg["is_junior"] = is_junior # 🌟 읽어온 막내 데이터를 cfg 바구니에 안전하게 담아서 보냄

        no_night = [False] * len(nurses)
        ws_r = wb_data["Restrictions"]
        for r in range(2, ws_r.max_row + 1):
            nm = ws_r.cell(r, 1).value
            if nm is None or str(nm).strip() == "":
                continue
            nm = str(nm).strip()
            if nm not in nurse_index:
                raise ValueError(f"Restrictions sheet nurse {nm!r} is not in Nurses sheet.")
            yn = normalize_text(ws_r.cell(r, 2).value)
            no_night[nurse_index[nm]] = yn == "Y"

        off_requests = [dict() for _ in nurses]
        prefs = [dict() for _ in nurses]

        ws_off = wb_data["OffRequests"]
        ws_pref = wb_data["ShiftPreferences"]

        for r in range(2, 2 + len(nurses)):
            nm = str(ws_off.cell(r, 1).value).strip()
            if nm != nurses[r - 2]:
                raise ValueError("OffRequests nurse order must match Nurses sheet.")
            nm2 = str(ws_pref.cell(r, 1).value).strip()
            if nm2 != nurses[r - 2]:
                raise ValueError("ShiftPreferences nurse order must match Nurses sheet.")

            for d in range(days):
                off_raw = ws_off.cell(r, d + 2).value
                pref_raw = normalize_text(ws_pref.cell(r, d + 2).value)

                if off_raw is not None and str(off_raw).strip() != "":
                    val = str(off_raw).strip().upper()
                    if val in ["1", "O", "OFF", "Y", "TRUE", "1순위"]:
                        off_requests[r - 2][d] = 1  
                    elif val in ["2", "2순위"]:
                        off_requests[r - 2][d] = 2  
                    elif val in ["3", "3순위"]:
                        off_requests[r - 2][d] = 3  
                    else:
                        off_requests[r - 2][d] = 1  

                if pref_raw:
                    clean_pref = pref_raw.replace("OR", " ").replace("/", " ").replace(",", " ")
                    pref_list = [s for s in clean_pref.split() if s in PREF_MARKERS]
                    if pref_list:
                        prefs[r - 2][d] = pref_list

        prev_shifts = {n: {} for n in range(len(nurses))}
        try:
            ws_prev = wb_data["PrevMonth"]
            for r in range(2, ws_prev.max_row + 1):
                nm = str(ws_prev.cell(r, 1).value).strip()
                if nm in nurse_index:
                    n_idx = nurse_index[nm]
                    for i in range(5):
                        val = normalize_text(ws_prev.cell(r, i + 2).value)
                        if val not in SHIFT_CODES:
                            val = "O"
                        prev_shifts[n_idx][-5 + i] = val
        except KeyError:
            for n in range(len(nurses)):
                for d in range(-5, 0):
                    prev_shifts[n][d] = "O"

        return wb, cfg, nurses, is_senior, is_nk, max_offs, holidays, no_night, off_requests, prefs, prev_shifts, preceptors, is_ghost

    def day_type(cfg, holidays, d):
        dt = date(cfg["year"], cfg["month"], d + 1)
        is_weekend_or_holiday = dt.weekday() >= 5 or d in holidays
        return "Weekend/Holiday" if is_weekend_or_holiday else "Weekday"

    def req_for_day(cfg, holidays, d):
        dt = day_type(cfg, holidays, d)
        if dt == "Weekend/Holiday":
            req = dict(cfg["weekend_req"])
        else:
            req = dict(cfg["weekday_req"])

        if "extra_reqs" in cfg and d in cfg["extra_reqs"]:
            for s, cnt in cfg["extra_reqs"][d].items():
                req[s] = req.get(s, 0) + cnt

        return req

    def preliminary_checks(cfg, nurses, holidays, no_night, off_requests):
        days = get_days_in_month(cfg["year"], cfg["month"])
        required_total = []
        issues = []

        for d in range(days):
            req = req_for_day(cfg, holidays, d)
            total_req = req["D"] + req["E"] + req["N"]
            required_total.append(total_req)

            off_cnt = sum(1 for n in range(len(nurses)) if d in off_requests[n] and off_requests[n][d] == 1)
            if len(nurses) - off_cnt < total_req:
                issues.append(
                    f"Day {d+1}: 1순위 오프 신청자가 너무 많아 {total_req}명의 근무자를 채울 수 없습니다."
                )

            night_avail = sum(
                1
                for n in range(len(nurses))
                if (d not in off_requests[n] or off_requests[n][d] > 1) and (not no_night[n])
            )
            if night_avail < req["N"]:
                issues.append(
                    f"Day {d+1}: 나이트 가능한 인원이 {night_avail}명뿐이라 필요 인원({req['N']}명)을 채울 수 채울 수 없습니다."
                )

        return issues

    def build_model(cfg, nurses, is_senior, is_nk, max_offs, holidays, no_night, off_requests, prefs, prev_shifts, preceptors, is_ghost):
        model = cp_model.CpModel()
        N = len(nurses)
        D = get_days_in_month(cfg["year"], cfg["month"])
        H = 5
        
        obj_terms = [] 
        x = {}

        for n in range(N):
            for d in range(-H, D):
                for s in SHIFT_CODES:
                    x[n, d, s] = model.NewBoolVar(f"x_{n}_{d}_{s}")
                    if d < 0:
                        past_s = prev_shifts[n].get(d, "O")
                        if s == past_s:
                            model.Add(x[n, d, s] == 1)
                        else:
                            model.Add(x[n, d, s] == 0)

        for n in range(N):
            for d in range(D):
                model.Add(sum(x[n, d, s] for s in SHIFT_CODES) == 1)
                if is_nk[n]:
                    model.Add(x[n, d, "D"] == 0)
                    model.Add(x[n, d, "E"] == 0)
                    if "M" in SHIFT_CODES:
                        model.Add(x[n, d, "M"] == 0)

        for d in range(D):
            req = req_for_day(cfg, holidays, d)
            for s in WORK_CODES:
                model.Add(sum(x[n, d, s] for n in range(N) if not is_ghost[n]) == req[s])
                
                if req[s] > 0:
                    senior_count = sum(x[n, d, s] for n in range(N) if s in is_senior[n] and not is_ghost[n])
                    
                    if s == "M":
                        obj_terms.append(500 * senior_count)
                    else:
                        missing_senior = model.NewBoolVar(f"missing_senior_{d}_{s}")
                        model.Add(missing_senior >= 1 - senior_count)
                        obj_terms.append(1000 * missing_senior)
                        
                        if req[s] >= 2:
                            ideal_max_senior = max(1, req[s] - 1)
                            extra_senior = model.NewIntVar(0, N, f"extra_senior_{d}_{s}")
                            model.Add(senior_count - ideal_max_senior <= extra_senior)
                            obj_terms.append(400 * extra_senior)

        for n in range(N):
            for d, rank in off_requests[n].items():
                if rank == 1:
                    model.Add(x[n, d, "O"] == 1)
                else:
                    penalty = 10000 if rank == 2 else 1000
                    miss_var = model.NewBoolVar(f"off_miss_{n}_{d}")
                    model.Add(x[n, d, "O"] + miss_var == 1)
                    obj_terms.append(penalty * miss_var)

        for n in range(N):
            if no_night[n]:
                for d in range(D):
                    model.Add(x[n, d, "N"] == 0)

        for n in range(N):
            if max_offs[n] < D:
                model.Add(sum(x[n, d, "O"] for d in range(D)) <= max_offs[n])

        for n in range(N):
            if preceptors[n]:
                for d in range(D):
                    for s in WORK_CODES:
                        model.Add(x[n, d, s] <= sum(x[p, d, s] for p in preceptors[n]))
                
                if is_ghost[n]:
                    min_work = D - max_offs[n] - 2
                    if min_work < 0: min_work = 0
                    model.Add(sum(x[n, d, s] for d in range(D) for s in WORK_CODES) >= min_work)

        for n in range(N):
            if is_nk[n]:
                model.Add(sum(x[n, d, "N"] for d in range(D)) == 15)
                
                nk_max_night = cfg["max_night_run"]
                if nk_max_night >= 0 and D >= nk_max_night + 1:
                    start_bound_max = max(-H, -nk_max_night)
                    for start in range(start_bound_max, D - nk_max_night):
                        model.Add(sum(x[n, d, "N"] for d in range(start, start + nk_max_night + 1)) <= nk_max_night)
                
                nk_min_night = cfg.get("min_night_run", 1)
                if nk_min_night > 1:
                    for d in range(D):
                        prev_N = x[n, d - 1, "N"] if d > -H else 0
                        for j in range(1, nk_min_night):
                            if d + j < D: model.Add(x[n, d, "N"] - prev_N <= x[n, d + j, "N"])
                        
                        if d < D - 1:
                            next_N = x[n, d + 1, "N"]
                            for j in range(1, nk_min_night):
                                if d - j >= -H: model.Add(x[n, d, "N"] - next_N <= x[n, d - j, "N"])

                post_off = cfg["post_night_off"]
                if post_off > 0:
                    for d in range(-1, D - 1):
                        end_block = model.NewBoolVar(f"nk_endnight_{n}_{d}")
                        model.Add(end_block <= x[n, d, "N"])
                        model.Add(end_block <= 1 - x[n, d + 1, "N"])
                        model.Add(end_block >= x[n, d, "N"] - x[n, d + 1, "N"])
                        for k in range(1, post_off + 1):
                            if d + k < D: model.Add(x[n, d + k, "O"] >= end_block)

        max_work_run = cfg["max_work_run"]
        if max_work_run >= 0 and D >= max_work_run + 1:
            for n in range(N):
                if is_nk[n]: continue
                start_bound = max(-H, -max_work_run)
                for start in range(start_bound, D - max_work_run):
                    model.Add(sum(x[n, d, s] for d in range(start, start + max_work_run + 1) for s in WORK_CODES) <= max_work_run)

        max_d = cfg.get("max_consecutive_d", 4)
        max_e = cfg.get("max_consecutive_e", 4)
        if max_d > 0 and D >= max_d + 1:
            for n in range(N):
                if is_nk[n]: continue
                start_bound = max(-H, -max_d)
                for d in range(start_bound, D - max_d):
                    model.Add(sum(x[n, d + i, "D"] for i in range(max_d + 1)) <= max_d)
        if max_e > 0 and D >= max_e + 1:
            for n in range(N):
                if is_nk[n]: continue
                start_bound = max(-H, -max_e)
                for d in range(start_bound, D - max_e):
                    model.Add(sum(x[n, d + i, "E"] for i in range(max_e + 1)) <= max_e)

        max_night_run = cfg["max_night_run"]
        if max_night_run >= 0 and D >= max_night_run + 1:
            for n in range(N):
                if is_nk[n]: continue
                start_bound = max(-H, -max_night_run)
                for start in range(start_bound, D - max_night_run):
                    model.Add(sum(x[n, d, "N"] for d in range(start, start + max_night_run + 1)) <= max_night_run)

        min_night_run = cfg.get("min_night_run", 1)
        if min_night_run > 1:
            for n in range(N):
                if is_nk[n]: continue
                for d in range(D):
                    prev_N = x[n, d - 1, "N"] if d > -H else 0
                    for j in range(1, min_night_run):
                        if d + j < D: model.Add(x[n, d, "N"] - prev_N <= x[n, d + j, "N"])
                    
                    if d < D - 1:
                        next_N = x[n, d + 1, "N"]
                        for j in range(1, min_night_run):
                            if d - j >= -H: model.Add(x[n, d, "N"] - next_N <= x[n, d - j, "N"])

        for n in range(N):
            if is_nk[n]: continue
            
            for d in range(-1, D - 1):
                model.Add(x[n, d, "E"] + x[n, d + 1, "D"] <= 1)
                if "M" in SHIFT_CODES:
                    model.Add(x[n, d, "E"] + x[n, d + 1, "M"] <= 1)  
                    model.Add(x[n, d, "M"] + x[n, d + 1, "D"] <= 1)  

            for d in range(-2, D - 2):
                model.Add(x[n, d, "O"] + sum(x[n, d + 1, s] for s in WORK_CODES) + x[n, d + 2, "O"] <= 2)
                model.Add(x[n, d, "N"] + x[n, d + 1, "O"] + x[n, d + 2, "D"] <= 2)

        post_off = cfg["post_night_off"]
        if post_off > 0:
            for n in range(N):
                if is_nk[n]: continue
                for d in range(-1, D - 1):
                    end_block = model.NewBoolVar(f"endnight_{n}_{d}")
                    model.Add(end_block <= x[n, d, "N"])
                    model.Add(end_block <= 1 - x[n, d + 1, "N"])
                    model.Add(end_block >= x[n, d, "N"] - x[n, d + 1, "N"])
                    for k in range(1, post_off + 1):
                        if d + k < D: model.Add(x[n, d + k, "O"] >= end_block)

        for n in range(N):
            if is_nk[n]: continue  
            for d in range(-H, D - 2):
                for gap in range(2, 10): 
                    if d + gap < D:
                        non_viol = model.NewBoolVar(f"non_viol_{n}_{d}_{gap}")
                        model.Add(
                            x[n, d, "N"] + x[n, d + gap, "N"] + 
                            sum(x[n, d + k, "O"] for k in range(1, gap)) - gap <= non_viol
                        )
                        obj_terms.append(500 * non_viol)

        pref_weight = cfg["pref_weight"]
        pref_miss = {}
        for n in range(N):
            for d, pref_list in prefs[n].items():
                miss = model.NewBoolVar(f"prefmiss_{n}_{d}")
                pref_miss[n, d] = miss
                model.Add(miss == 1 - sum(x[n, d, p] for p in pref_list))
                obj_terms.append(pref_weight * miss)

        # ====================================================================
        # 🌟 [추가 로직] 막내 간호사 동시 근무 방지 (벌점 시스템)
        # ====================================================================
        is_junior_list = cfg.get("is_junior", [])
        if is_junior_list:
            junior_nurses = [i for i, jun in enumerate(is_junior_list) if jun == 'Y']
            if len(junior_nurses) > 0:
                JUNIOR_OVERLAP_PENALTY = 5000
                for d in range(D):
                    for s in WORK_CODES: # D, E, N, M 모든 근무를 감시
                        current_juniors = sum(x[n, d, s] for n in junior_nurses)
                        excess_juniors = model.NewIntVar(0, len(junior_nurses), f'excess_juniors_d{d}_{s}')
                        model.Add(current_juniors - 1 <= excess_juniors)
                        obj_terms.append(excess_juniors * JUNIOR_OVERLAP_PENALTY)
        # ====================================================================

        weekend_days = [d for d in range(D) if day_type(cfg, holidays, d) == "Weekend/Holiday"]
        
        total_work_vars = []
        total_night_vars = []
        total_day_vars = []
        total_eve_vars = []
        total_weekend_vars = []

        for n in range(N):
            tw = model.NewIntVar(0, D, f"totwork_{n}")
            tn = model.NewIntVar(0, D, f"totnight_{n}")
            td = model.NewIntVar(0, D, f"totday_{n}")
            te = model.NewIntVar(0, D, f"toteve_{n}")
            twk = model.NewIntVar(0, len(weekend_days), f"totweekend_{n}")

            model.Add(tw == sum(x[n, d, s] for d in range(D) for s in WORK_CODES))
            model.Add(tn == sum(x[n, d, "N"] for d in range(D)))
            model.Add(td == sum(x[n, d, "D"] for d in range(D)))
            model.Add(te == sum(x[n, d, "E"] for d in range(D)))
            model.Add(twk == sum(x[n, d, s] for d in weekend_days for s in WORK_CODES))

            total_work_vars.append(tw)
            total_night_vars.append(tn)
            total_day_vars.append(td)
            total_eve_vars.append(te)
            total_weekend_vars.append(twk)

        normal_nurses = [n for n in range(N) if not is_nk[n] and not preceptors[n]]
        
        if normal_nurses:
            max_tot = model.NewIntVar(0, D, "max_tot")
            min_tot = model.NewIntVar(0, D, "min_tot")
            model.AddMaxEquality(max_tot, [total_work_vars[n] for n in normal_nurses])
            model.AddMinEquality(min_tot, [total_work_vars[n] for n in normal_nurses])
            obj_terms.append(cfg["fair_total_weight"] * (max_tot - min_tot))

            def add_group_fairness(group, vars_list, weight, prefix):
                if len(group) <= 1 or weight <= 0: return
                max_v = model.NewIntVar(0, D, f"max_{prefix}")
                min_v = model.NewIntVar(0, D, f"min_{prefix}")
                model.AddMaxEquality(max_v, [vars_list[i] for i in group])
                model.AddMinEquality(min_v, [vars_list[i] for i in group])
                obj_terms.append(weight * (max_v - min_v))

            sn_n_group = [n for n in normal_nurses if "N" in is_senior[n] and not no_night[n]]
            jn_n_group = [n for n in normal_nurses if "N" not in is_senior[n] and not no_night[n]]
            add_group_fairness(sn_n_group, total_night_vars, cfg["fair_night_weight"], "sn")
            add_group_fairness(jn_n_group, total_night_vars, cfg["fair_night_weight"], "jn")

            sn_d_group = [n for n in normal_nurses if "D" in is_senior[n]]
            jn_d_group = [n for n in normal_nurses if "D" not in is_senior[n]]
            add_group_fairness(sn_d_group, total_day_vars, cfg.get("fair_day_weight", 0), "sd")
            add_group_fairness(jn_d_group, total_day_vars, cfg.get("fair_day_weight", 0), "jd")

            sn_e_group = [n for n in normal_nurses if "E" in is_senior[n]]
            jn_e_group = [n for n in normal_nurses if "E" not in is_senior[n]]
            add_group_fairness(sn_e_group, total_eve_vars, cfg.get("fair_evening_weight", 0), "se")
            add_group_fairness(jn_e_group, total_eve_vars, cfg.get("fair_evening_weight", 0), "je")

            sn_wk_group = [n for n in normal_nurses if is_senior[n]]
            jn_wk_group = [n for n in normal_nurses if not is_senior[n]]
            add_group_fairness(sn_wk_group, total_weekend_vars, cfg.get("fair_weekend_weight", 0), "sw")
            add_group_fairness(jn_wk_group, total_weekend_vars, cfg.get("fair_weekend_weight", 0), "jw")

        model.Minimize(sum(obj_terms))
        return model, x, pref_miss

    def solve_model(model, cfg):
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = cfg["time_limit"]
        solver.parameters.num_search_workers = 8
        solver.parameters.log_search_progress = False
        status = solver.Solve(model)
        return solver, status

    def compute_runs(seq, code_set):
        max_run = 0
        cur = 0
        for x in seq:
            if x in code_set:
                cur += 1
                max_run = max(max_run, cur)
            else:
                cur = 0
        return max_run

    def validate_schedule(schedule, nurses, cfg, holidays, no_night, off_requests, is_ghost):
        D = len(schedule[0])
        nurse_rows = []
        daily_rows = []

        for n, nm in enumerate(nurses):
            seq = schedule[n]
            max_work = compute_runs(seq, set(WORK_CODES))
            max_night = compute_runs(seq, {"N"})

            night_rest_viol = 0
            for d in range(D - 1):
                if seq[d] == "N" and seq[d + 1] != "N":
                    for k in range(1, cfg["post_night_off"] + 1):
                        if d + k < D and seq[d + k] != "O":
                            night_rest_viol += 1

            e_to_d_viol = sum(1 for d in range(D - 1) if seq[d] == "E" and seq[d + 1] == "D")
            pongdang_viol = sum(
                1 for d in range(D - 2) if seq[d] == "O" and seq[d + 1] in WORK_CODES and seq[d + 2] == "O"
            )
            no_night_viol = 0
            if no_night[n]:
                no_night_viol = sum(1 for d in range(D) if seq[d] == "N")
            
            off_req_viol = sum(1 for d, rank in off_requests[n].items() if rank == 1 and seq[d] != "O")

            overall = "OK"
            if (max_work > cfg["max_work_run"] or max_night > cfg["max_night_run"] or night_rest_viol > 0
                or e_to_d_viol > 0 or pongdang_viol > 0 or no_night_viol > 0 or off_req_viol > 0):
                overall = "NG"

            nurse_rows.append([nm, max_work, max_night, night_rest_viol, e_to_d_viol, pongdang_viol, no_night_viol, off_req_viol, overall])

        for d in range(D):
            req = req_for_day(cfg, holidays, d)
            actual = {s: sum(1 for n in range(len(nurses)) if schedule[n][d] == s and not is_ghost[n]) for s in WORK_CODES}
            ok = "OK" if all(actual[s] == req[s] for s in WORK_CODES) else "NG"
            daily_rows.append([date(cfg["year"], cfg["month"], d + 1).isoformat(), day_type(cfg, holidays, d),
                               req.get("D", 0), actual.get("D", 0), req.get("E", 0), actual.get("E", 0), req.get("N", 0), actual.get("N", 0), ok])

        return nurse_rows, daily_rows

    def write_outputs(wb, cfg, nurses, holidays, no_night, off_requests, prefs, schedule, pref_results, is_ghost):
        D = len(schedule[0])

        ws = wb["Schedule"]
        ws.delete_rows(4, ws.max_row)

        for d in range(1, 32):
            ws.cell(2, d + 1).value = ""
            ws.cell(3, d + 1).value = d if d <= D else ""
        for d in range(D):
            ws.cell(2, d + 2).value = "WH" if day_type(cfg, holidays, d) == "Weekend/Holiday" else "WD"

        unmet_rows = []
        for n, nm in enumerate(nurses):
            seq = list(schedule[n])
            off_positions = [i for i, s in enumerate(seq) if s == "O"]
            
            base_offs = cfg.get("base_offs", 11)
            for idx, d in enumerate(off_positions):
                if idx >= base_offs:
                    seq[d] = "S"

            row = 4 + n
            ws.cell(row, 1).value = nm

            pref_req = len(prefs[n])
            pref_hit = 0
            pref_miss = 0

            for d in range(D):
                cell = ws.cell(row, d + 2)
                cell.value = seq[d]

                if schedule[n][d] == "O": cell.fill = GRAY_FILL
                elif seq[d] == "S": cell.fill = BLUE_FILL
                elif schedule[n][d] == "N": cell.fill = PINK_FILL

                if d in prefs[n]:
                    requested_list = prefs[n][d]
                    assigned = schedule[n][d]
                    req_str = " or ".join(requested_list)
                    
                    if assigned in requested_list:
                        pref_hit += 1
                        cell.comment = Comment(f"Preferred {req_str} — matched", "Scheduler")
                        cell.fill = GREEN_FILL
                    else:
                        pref_miss += 1
                        cell.comment = Comment(f"Preferred {req_str} — not matched (assigned {assigned})", "Scheduler")
                        cell.fill = YELLOW_FILL
                        unmet_rows.append([nm, date(cfg["year"], cfg["month"], d + 1).isoformat(), "ShiftPreference", req_str, assigned, "Unmet"])

                if d in off_requests[n]:
                    rank = off_requests[n][d]
                    if rank == 1:
                        cell.comment = Comment("1순위 오프 (필수 확정)", "Scheduler")
                        cell.fill = GRAY_FILL
                    else:
                        if schedule[n][d] == "O" or schedule[n][d] == "S":
                            cell.comment = Comment(f"{rank}순위 오프 (승인됨!)", "Scheduler")
                            cell.fill = GRAY_FILL
                        else:
                            cell.comment = Comment(f"🚨 {rank}순위 오프 반려 (인원 부족으로 {schedule[n][d]} 배정)", "Scheduler")
                            cell.fill = RED_FILL
                            unmet_rows.append([nm, date(cfg["year"], cfg["month"], d + 1).isoformat(), "OffRequest", f"{rank}순위 오프", schedule[n][d], "Denied"])

            work_count = sum(1 for s in schedule[n] if s in WORK_CODES)
            d_count = sum(1 for s in schedule[n] if s == "D")
            e_count = sum(1 for s in schedule[n] if s == "E")
            n_count = sum(1 for s in schedule[n] if s == "N")
            off_count = sum(1 for s in seq if s in {"O", "S"})
            s_count = sum(1 for s in seq if s == "S")
            weekend_work = sum(1 for d in range(D) if day_type(cfg, holidays, d) == "Weekend/Holiday" and schedule[n][d] in WORK_CODES)

            vals = [work_count, d_count, e_count, n_count, off_count, s_count, weekend_work, pref_req, pref_hit, pref_miss]
            for j, v in enumerate(vals, start=33):
                ws.cell(row, j).value = v

        for d in range(D + 1, 32):
            for r in range(4, 4 + len(nurses)):
                ws.cell(r, d + 1).value = ""

        rs = wb["RequestsSummary"]
        rs.delete_rows(2, rs.max_row)
        for i, rowvals in enumerate(unmet_rows, start=2):
            for j, v in enumerate(rowvals, start=1):
                rs.cell(i, j).value = v

        val_nurse_rows, val_daily_rows = validate_schedule(schedule, nurses, cfg, holidays, no_night, off_requests, is_ghost)

        vs = wb["Validation"]
        vs.delete_rows(2, vs.max_row)
        for i, rowvals in enumerate(val_nurse_rows, start=2):
            for j, v in enumerate(rowvals, start=1):
                cell = vs.cell(i, j)
                cell.value = v
                if j == 9: cell.fill = GREEN_FILL if v == "OK" else RED_FILL

        ds = wb["DailyCoverage"]
        ds.delete_rows(2, ds.max_row)
        for i, rowvals in enumerate(val_daily_rows, start=2):
            for j, v in enumerate(rowvals, start=1):
                cell = ds.cell(i, j)
                cell.value = v
                if j == 9: cell.fill = GREEN_FILL if v == "OK" else RED_FILL

        ws_r = wb["Restrictions"]
        for i, nm in enumerate(nurses, start=2):
            ws_r.cell(i, 1).value = nm

        for title in ["OffRequests", "ShiftPreferences", "Schedule"]:
            w = wb[title]
            for d in range(1, 32):
                w.cell(3 if title == "Schedule" else 1, d + 1).value = d if d <= D else ""

        wb["Setup"]["D2"] = "Run: python nurse_scheduler_complete.py nurse_scheduler_complete.xlsx"
        wb["Setup"]["D3"] = "Needs: pip install openpyxl ortools"

        return wb

    def extract_schedule(solver, x, nurses, cfg):
        N = len(nurses)
        D = get_days_in_month(cfg["year"], cfg["month"])
        schedule = []
        for n in range(N):
            row = []
            for d in range(D):
                assigned = None
                for s in SHIFT_CODES:
                    if solver.Value(x[n, d, s]):
                        assigned = s
                        break
                if assigned is None:
                    raise RuntimeError(f"No shift found for nurse {n} day {d}.")
                row.append(assigned)
            schedule.append(row)
        return schedule

    def main():
        if len(sys.argv) == 2:
            path = sys.argv[1]
        else:
            path = "nurse_scheduler_complete.xlsx"

        if not os.path.exists(path):
            print(f"\n[에러] '{path}' 파일을 찾을 수 없습니다!")
            print("엑셀 파일이 프로그램(.exe)과 같은 폴더에 있는지 확인해 주세요.")
            input("\n종료하려면 엔터를 누르세요...")
            sys.exit(1)

        wb, cfg, nurses, is_senior, is_nk, max_offs, holidays, no_night, off_requests, prefs, prev_shifts, preceptors, is_ghost = load_inputs(path)

        precheck_issues = preliminary_checks(cfg, nurses, holidays, no_night, off_requests)
        if precheck_issues:
            msg = "\n".join(precheck_issues)
            raise RuntimeError("Pre-check found infeasible inputs before solving:\n" + msg)

        model, x, pref_miss = build_model(cfg, nurses, is_senior, is_nk, max_offs, holidays, no_night, off_requests, prefs, prev_shifts, preceptors, is_ghost)
        solver, status = solve_model(model, cfg)

        if status == cp_model.INFEASIBLE:
            raise RuntimeError(
                "\n[에러] 현재 설정된 규칙들끼리 서로 충돌하여 근무표를 완성할 수 없습니다.\n"
                "▶ 해결 팁 1: Setup 시트에서 D, E, N 인원수를 조금 줄여보세요.\n"
                "▶ 해결 팁 2: 간호사들의 '1순위 오프 신청'이나 나이트 금지가 너무 많지 않은지 확인해 주세요."
            )
        elif status == cp_model.UNKNOWN:
            raise RuntimeError(
                "\n[시간 초과] 컴퓨터가 정답을 찾기엔 규칙이 너무 복잡해서 제한 시간 내에 풀지 못했습니다.\n"
                "▶ 해결 팁 1: Setup 시트에서 'Solver time limit (sec)'을 300초 이상으로 충분히 늘려주세요.\n"
                "▶ 해결 팁 2: 연속 근무 제한 숫자를 살짝 늘려주세요."
            )
        elif status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            raise RuntimeError(f"Solver failed with status: {solver.status_name(status)}")

        schedule = extract_schedule(solver, x, nurses, cfg)
        wb = write_outputs(wb, cfg, nurses, holidays, no_night, off_requests, prefs, schedule, pref_miss, is_ghost)
        
        try:
            wb.save(path)
            print(f"\n✅ 성공! 근무표가 '{path}' 파일에 저장되었습니다.")
        except PermissionError:
            raise RuntimeError(f"\n[에러] '{path}' 엑셀 파일이 현재 켜져있습니다! 엑셀을 완전히 닫고 다시 실행해 주세요.")
            
        print(f"Solver status: {solver.status_name(status)}")
        print(f"Objective value: {solver.ObjectiveValue()}")
        
        input("\n🎉 스케줄러 작성이 완료되었습니다! 확인하셨으면 엔터를 눌러 창을 닫아주세요...")

    if __name__ == "__main__":
        main()

except Exception as e:
    print("\n========================================================")
    print("[치명적인 오류 발생] 프로그램 시작 중 문제가 발생했습니다!")
    print("========================================================")
    traceback.print_exc()
    print("========================================================")
    input("\n위 에러 메시지를 사진으로 찍거나 복사한 뒤 엔터를 눌러 종료하세요...")